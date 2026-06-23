"""Из markdown-таблицы сравнения конкурентов собрать .xlsx (для заливки как Google Sheet).

Бот в анализе конкурентов уже выдаёт сравнительную таблицу в markdown в нужной
раскладке: первый столбец — параметры, дальше по столбцу на конкурента (как в
образце «Бесплатники моб разраб.xlsx»). Здесь её парсим и кладём в .xlsx с лёгким
оформлением. Дальше drive_upload.upload_as_gsheet превращает .xlsx в Google Таблицу.
"""
from __future__ import annotations

import logging
import re
from pathlib import Path

log = logging.getLogger("pult-bot.sheet")

# Жирный markdown (**текст**) и инлайн-код — чистим, в ячейке не нужны.
_BOLD = re.compile(r"\*\*(.+?)\*\*")
_SEP_ROW = re.compile(r"^\s*\|?[\s:|-]+\|?\s*$")  # строка-разделитель |---|---|


def _clean(cell: str) -> str:
    cell = _BOLD.sub(r"\1", cell).strip()
    return cell.strip("`").strip()


def extract_table(md_text: str) -> list[list[str]] | None:
    """Достать ПЕРВУЮ (самую крупную) markdown-таблицу как список строк-списков.
    Вернёт None, если таблицы нет."""
    blocks: list[list[list[str]]] = []
    cur: list[list[str]] = []
    for line in md_text.splitlines():
        s = line.strip()
        if s.startswith("|") and s.count("|") >= 2:
            if _SEP_ROW.match(s):
                continue  # пропускаем |---|---|
            cells = [_clean(c) for c in s.strip("|").split("|")]
            cur.append(cells)
        else:
            if len(cur) >= 2:
                blocks.append(cur)
            cur = []
    if len(cur) >= 2:
        blocks.append(cur)
    if not blocks:
        return None
    # самая большая таблица = матрица сравнения
    return max(blocks, key=lambda b: len(b) * max(len(r) for r in b))


def build_xlsx(rows: list[list[str]], xlsx_path: Path, sheet_title: str = "Сравнение") -> Path | None:
    """Записать строки в .xlsx с лёгким оформлением. Вернуть путь или None."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.warning("openpyxl не установлен — Google-таблица не собрана.")
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31] or "Сравнение"
    ncols = max(len(r) for r in rows)

    head_fill = PatternFill("solid", fgColor="D9E1F2")
    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    for ri, row in enumerate(rows, start=1):
        for ci in range(ncols):
            val = row[ci] if ci < len(row) else ""
            cell = ws.cell(row=ri, column=ci + 1, value=val)
            cell.alignment = wrap
            if ri == 1 or ci == 0:          # шапка-строка и первый столбец (параметры)
                cell.font = bold
                if ri == 1:
                    cell.fill = head_fill

    # ширины: первый столбец пошире (названия параметров), остальные ровно
    ws.column_dimensions["A"].width = 26
    for ci in range(2, ncols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 34
    ws.freeze_panes = "B2"  # шапка + первый столбец закреплены

    try:
        wb.save(str(xlsx_path))
        return xlsx_path
    except Exception:
        log.exception("Не удалось сохранить .xlsx")
        return None


def md_to_xlsx(md_text: str, xlsx_path: Path, sheet_title: str = "Сравнение") -> Path | None:
    """Полный путь: markdown-текст → .xlsx с таблицей сравнения. None, если таблицы нет."""
    rows = extract_table(md_text)
    if not rows:
        log.info("В результате анализа не нашлось markdown-таблицы — Google-таблицу не делаю.")
        return None
    return build_xlsx(rows, xlsx_path, sheet_title)
